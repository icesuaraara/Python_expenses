import openpyxl
import os
import pandas as pd
from datetime import datetime

def creatfile():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'รายการของที่ซื้อ'
    sheet['B1'] = 'ราคา'
    sheet['C1'] = 'เวลา'
    workbook.save('รายจ่าย.xlsx')

def checkfile():
    current_directory = os.getcwd() 
    folder_path = current_directory 
    file_list = os.listdir(folder_path)
    for filename in file_list:
        if "รายจ่าย.xlsx" in file_list:
            print("\nCheckfile succeed\n")
            break
        else:
            confirm = False
            checkcreatfile = str(input("\nYou don't have an Excel file.\n"+"Will you create a new file? [y/n] : "))
            if (checkcreatfile == "y") or (checkcreatfile == "Y"):
                confirm = True
            elif (checkcreatfile == "n") or (checkcreatfile == "N"):
                confirm = False
                print("\n-----Exit Program-----\n")
            else:
                print("Try again")
            if confirm:
                creatfile()
                print("\nCreatfile succeed\n") 
def menu():
    print("-"*5,"Progrram Expenses","-"*5)
    print("\n1. ดูรายการ")
    print("2. เพิ่มรายการ")
    print("3. ดูยอด")
    print("4. ออกจากโปรแกรม")
    while True:
        menu_select = int(input("\nเลือกเมนูใช้งานโปรแกรม : "))
        if menu_select == 1:
            openfile()
        elif menu_select == 2:
            append_data()
        elif menu_select == 3:
            sumprice()
        elif menu_select == 4:
            print("\n-----Exit Program-----\n")
            break
        else:
            print("\n-----Try again-----\n")
def openfile():
    print("Open file")
    df = pd.read_excel('รายจ่าย.xlsx', sheet_name='Sheet')
    print(df)
    
def append_data():
    print("")
    workbook = openpyxl.load_workbook('รายจ่าย.xlsx')
    sheet = workbook['Sheet']

    add_list_data = str(input("เพิ่มรายการของที่ซื้อ: "))
    add_price_data = int(input("ราคาของที่ซื้อ: "))
    column_list = 'A'
    column_price = 'B'
    column_time = 'C'
    row_count = len([cell.value for cell in sheet[column_list]]) #นับจำนวนแถวคอลัม A
    sheet[column_list + str(row_count+1)] = add_list_data #เลือกcolumn_letterแล้วเลือกแถวแล้วเพิ่มข้อมูล
    sheet[column_price+ str(row_count+1)] = add_price_data
    current_datetime = datetime.now()
    sheet[column_time + str(row_count+1)] = current_datetime
    workbook.save('รายจ่าย.xlsx')
    workbook.close()
def sumprice():
    print("")
    workbook = openpyxl.load_workbook('รายจ่าย.xlsx')
    sheet = workbook['Sheet']
    column_letter = 'B'
    column_data = []
    for cell in sheet[column_letter]:
        column_data.append(cell.value)
    workbook.close()
    cal_sum = column_data[1:]
    sum_data = 0
    for i in range(len(cal_sum)):
        sum_data += int(cal_sum[i])
    print("ราคารวมของที่ซื้อมาทั้งหมด :",sum_data)

checkfile()
menu()